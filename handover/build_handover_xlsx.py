"""Builds the school data-handover Excel workbook.

One workbook, one sheet per data category. Sheet 1 (README) explains what is
needed, why, when, and how to fill in. Subsequent sheets are templates with
header rows + 1-2 sample rows + data validation drop-downs where applicable.

Run from anywhere; output → handover/Manhaj_Data_Handover_Template.xlsx
"""
from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

ROOT = Path(__file__).resolve().parent
OUT  = ROOT / "Manhaj_Data_Handover_Template.xlsx"

NAVY     = "0B2545"
ACCENT   = "3D5A80"
SOFT     = "EEF2F7"
WHITE    = "FFFFFF"
MUTED    = "6B7C93"
WARN     = "FFF8E1"
INK      = "1A2440"

thin = Side(border_style="thin", color="E5EAF0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(cell):
    cell.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    cell.fill = PatternFill(fill_type="solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = BORDER

def style_sub(cell):
    cell.font = Font(name="Calibri", bold=True, color=INK, size=10)
    cell.fill = PatternFill(fill_type="solid", fgColor=SOFT)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = BORDER

def style_body(cell):
    cell.font = Font(name="Calibri", color=INK, size=10)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = BORDER

def make_readme(ws):
    ws.title = "README"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 90

    rows = [
        ("Manhaj data handover template",           ""),
        ("School",                                  "International School of Oman"),
        ("Academic year",                           "2026 / 2027"),
        ("Prepared for",                            "School administration"),
        ("",                                        ""),
        ("Why we're asking",                        ("Manhaj is a unified Admin + Classroom + Finance platform. To demo the parts of the platform that need school-specific data, please fill in the sheets below. The fields are split into REQUIRED (needed for V1) and OPTIONAL (nicer demo, but not blocking). If a field is sensitive, leave it blank and tag the cell with a note — we'll discuss handling.")),
        ("How to use this file",                    "Each sheet covers one data category. Read the orange banner on each sheet — it says what's required, what's optional, and how columns are used. Sample rows are pre-filled so you can see the shape. Save the file and email back to the Manhaj team."),
        ("Confidentiality",                         "We process data on your behalf only. Sensitive sheets (Finance, SEN) are accessed by Manhaj engineering on a need-to-know basis. Region applicable: PDPL Oman."),
        ("",                                        ""),
        ("Sheet",                                   "Purpose · what it unlocks"),
        ("1. Students (REQUIRED)",                  "Roster: who's in the school next year. Unlocks the entire student-facing layer (course selection, attendance, parent reports)."),
        ("2. Parents (REQUIRED)",                   "Contact channels for digital course-selection and monthly reports. WhatsApp opt-in flag is critical."),
        ("3. Bell schedule (REQUIRED)",             "When the school day starts and ends and how many periods. Powers the timetable engine."),
        ("4. Rooms (REQUIRED minimal)",             "Where classes can run. Capacity + equipment flags. Demo can use a 'default 30-cap classroom' if unavailable."),
        ("5. Term calendar (REQUIRED)",             "Start/end dates per term + public holidays. Drives attendance + report cadence."),
        ("6. Lesson plans (OPTIONAL)",              "Subject + objectives per period. Optional for demo; powers the classroom recap features once available."),
        ("7. Assessments (OPTIONAL)",               "Past assessment scores. Optional for demo. If unavailable, we use a Manhaj-generated illustrative dataset and clearly label it."),
        ("8. Chart of accounts (OPTIONAL)",         "Finance module only. Send when the Finance pilot conversation starts. NOT needed for the initial Admin + Classroom demo."),
        ("9. Alumni placements (OPTIONAL)",         "Last 5 years of university placements. Unlocks the university-fit signal in monthly reports. We can demo without — but quality jumps with real data."),
        ("",                                        ""),
        ("What Manhaj provides (no input needed from you)", "Rubric framework (the 6 universal axes), report-card and monthly-report templates, parent communication templates, university-fit logic. These are part of the product — schools tune descriptors, not the framework."),
        ("",                                        ""),
        ("Deadline",                                "Please return REQUIRED sheets by — TBC with Manhaj team."),
        ("Questions",                               "Email the Manhaj team or reply to the WhatsApp thread."),
    ]
    for i, (label, value) in enumerate(rows, start=1):
        a = ws.cell(row=i, column=1, value=label)
        b = ws.cell(row=i, column=2, value=value)
        if i == 1:
            a.font = Font(name="Calibri", bold=True, color=WHITE, size=18)
            a.fill = PatternFill(fill_type="solid", fgColor=NAVY)
            b.fill = PatternFill(fill_type="solid", fgColor=NAVY)
            a.alignment = Alignment(vertical="center")
            ws.row_dimensions[i].height = 36
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
        elif i == 10:
            style_sub(a); style_sub(b)
        elif label.startswith(("1.", "2.", "3.", "4.", "5.", "6.", "7.", "8.", "9.")):
            a.font = Font(name="Calibri", bold=True, size=10, color=INK)
            b.font = Font(name="Calibri", size=10, color=INK)
            a.alignment = Alignment(vertical="top", wrap_text=True)
            b.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[i].height = 32
        elif label and not value:
            a.font = Font(name="Calibri", bold=True, size=11, color=NAVY)
        else:
            a.font = Font(name="Calibri", bold=True, size=10, color=INK)
            b.font = Font(name="Calibri", size=10, color=INK)
            a.alignment = Alignment(vertical="top", wrap_text=True)
            b.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[i].height = 28


def add_banner(ws, text, color=WARN):
    ws.merge_cells("A1:Z1")
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(name="Calibri", bold=True, color="744210", size=10)
    c.fill = PatternFill(fill_type="solid", fgColor=color)
    c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
    ws.row_dimensions[1].height = 36


def add_template(wb, title, banner, columns, samples, dropdowns=None):
    """columns = [(header, width, note_or_None)]. samples = list[dict|list]."""
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    add_banner(ws, banner)
    # Headers
    for i, (hdr, width, note) in enumerate(columns, start=1):
        c = ws.cell(row=2, column=i, value=hdr)
        style_header(c)
        ws.column_dimensions[get_column_letter(i)].width = width
        if note:
            c.comment = Comment(note, "Manhaj")
    # Samples
    for r, sample in enumerate(samples, start=3):
        for i, val in enumerate(sample, start=1):
            c = ws.cell(row=r, column=i, value=val)
            style_body(c)
            c.font = Font(name="Calibri", color=MUTED, size=10, italic=True)

    # Dropdowns
    if dropdowns:
        for col_idx, options in dropdowns.items():
            dv = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=True)
            dv.add(f"{get_column_letter(col_idx)}3:{get_column_letter(col_idx)}1000")
            ws.add_data_validation(dv)

    ws.freeze_panes = "A3"
    return ws


def main():
    wb = Workbook()
    make_readme(wb.active)

    # ---- 1. STUDENTS (REQUIRED) ----
    add_template(wb, "1. Students (REQUIRED)",
        "REQUIRED · Roster of students enrolled for AY 2026/27. One row per student. "
        "Internal student IDs accepted in any format — we'll preserve them as your external_ref.",
        columns=[
            ("Student ID (school's own)", 22, "Free-form. We use this as the link key to all your existing systems."),
            ("Full name (English)", 28, "As it should appear in reports."),
            ("Full name (Arabic)", 28, "Optional but recommended for bilingual reports."),
            ("Date of birth", 14, "YYYY-MM-DD"),
            ("Gender", 12, "M / F"),
            ("Nationality", 18, ""),
            ("Grade level 2026/27", 14, "KG1, KG2, 1..12"),
            ("Section code", 14, "Match the section codes from the 26-27A planning sheet, e.g. 9A, 11 AS, 12 A2"),
            ("Enrolled date", 14, "YYYY-MM-DD · when student joined the school"),
            ("SEN flag", 10, "Y / N — leave blank if not applicable"),
            ("Primary parent ID", 16, "Links to the Parents sheet"),
            ("Secondary parent ID", 16, "Optional · links to the Parents sheet"),
            ("Notes", 36, "Anything we should know"),
        ],
        samples=[
            ["ISO-2843", "Layla Al-Habsi", "ليلى الحبسي", "2010-04-12", "F", "Omani", "10", "10A", "2017-09-01", "N", "P-1001", "P-1002", ""],
            ["ISO-2855", "Khalil Al-Lawati", "خليل اللواتي", "2011-08-30", "M", "Omani", "9", "9B", "2018-09-05", "N", "P-1003", "", ""],
        ],
        dropdowns={5: ["M", "F"], 7: ["KG1","KG2","1","2","3","4","5","6","7","8","9","10","11","12"], 10: ["Y","N"]},
    )

    # ---- 2. PARENTS (REQUIRED) ----
    add_template(wb, "2. Parents (REQUIRED)",
        "REQUIRED · One row per parent / guardian. WhatsApp opt-in is critical for parent comms and course-selection links.",
        columns=[
            ("Parent ID (school's own)", 18, "Link key — appears in the Students sheet."),
            ("Full name", 28, ""),
            ("Relationship", 16, ""),
            ("Phone (E.164)", 18, "Start with +, no spaces. e.g. +96891234567"),
            ("Email", 28, ""),
            ("WhatsApp opt-in", 16, "Y / N — must be Y for digital course-selection and report delivery"),
            ("Preferred language", 18, "EN / AR"),
            ("Notes", 30, ""),
        ],
        samples=[
            ["P-1001", "Saud Al-Habsi", "Father", "+96891000001", "saud.h@example.com", "Y", "EN", "Primary contact"],
            ["P-1002", "Mona Al-Habsi", "Mother", "+96891000002", "mona.h@example.com", "Y", "AR", ""],
            ["P-1003", "Ahmed Al-Lawati", "Father", "+96891000003", "ahmed.l@example.com", "Y", "EN", ""],
        ],
        dropdowns={3: ["Father","Mother","Guardian"], 6: ["Y","N"], 7: ["EN","AR"]},
    )

    # ---- 3. BELL SCHEDULE (REQUIRED) ----
    add_template(wb, "3. Bell schedule (REQUIRED)",
        "REQUIRED · The day's bell schedule. One row per period per day type. "
        "If all days share the same schedule, set Day = 'All'.",
        columns=[
            ("Day", 14, "Monday..Sunday, or 'All' if same every day"),
            ("Period #", 12, "1, 2, 3..."),
            ("Period label", 18, "Optional — e.g. 'P1', 'Recess', 'Assembly'"),
            ("Start time", 14, "HH:MM (24-hour)"),
            ("End time", 14, "HH:MM (24-hour)"),
            ("Is teaching period", 18, "Y / N — recess / lunch / assembly = N"),
            ("Notes", 30, ""),
        ],
        samples=[
            ["All", 1, "P1", "07:45", "08:30", "Y", "First period"],
            ["All", 2, "P2", "08:30", "09:15", "Y", ""],
            ["All", 3, "Recess", "09:15", "09:35", "N", ""],
            ["All", 4, "P3", "09:35", "10:20", "Y", ""],
            ["All", 5, "P4", "10:20", "11:05", "Y", ""],
            ["All", 6, "Lunch", "11:05", "11:35", "N", ""],
            ["All", 7, "P5", "11:35", "12:20", "Y", ""],
            ["All", 8, "P6", "12:20", "13:05", "Y", ""],
            ["All", 9, "P7", "13:05", "13:50", "Y", ""],
        ],
        dropdowns={1: ["All","Sunday","Monday","Tuesday","Wednesday","Thursday","Friday","Saturday"], 6: ["Y","N"]},
    )

    # ---- 4. ROOMS (REQUIRED minimal) ----
    add_template(wb, "4. Rooms (REQUIRED min)",
        "REQUIRED (minimal) · Room inventory. If unavailable, we can stub with 'default 30-cap classrooms' for demo, "
        "but accurate room data unlocks the substitution + scheduling engine.",
        columns=[
            ("Room code", 16, "School's own room identifier"),
            ("Room name", 24, ""),
            ("Capacity", 12, "Max students"),
            ("Type", 16, "classroom / lab / gym / music / library / other"),
            ("Floor / building", 18, ""),
            ("Equipment (comma-separated)", 36, "e.g. 'smartboard, projector, sink'"),
            ("Available for scheduling", 22, "Y / N — N for offices, storage, etc"),
        ],
        samples=[
            ["R-101", "Year 9A Homeroom", 28, "classroom", "Block A · 1st floor", "smartboard, projector", "Y"],
            ["R-LAB1", "Science Lab 1", 24, "lab", "Block B · ground floor", "smartboard, sinks, fume hood", "Y"],
            ["R-GYM", "Main Gym", 60, "gym", "Sports building", "PE equipment store", "Y"],
        ],
        dropdowns={4: ["classroom","lab","gym","music","library","art","other"], 7: ["Y","N"]},
    )

    # ---- 5. TERM CALENDAR (REQUIRED) ----
    add_template(wb, "5. Term calendar (REQUIRED)",
        "REQUIRED · Academic terms + public holidays + school breaks for AY 2026/27.",
        columns=[
            ("Type", 14, "term / holiday / break / exam-week"),
            ("Label", 28, ""),
            ("Start date", 14, "YYYY-MM-DD"),
            ("End date", 14, "YYYY-MM-DD"),
            ("Notes", 30, ""),
        ],
        samples=[
            ["term", "Term 1", "2026-09-01", "2026-12-18", ""],
            ["holiday", "Eid al-Adha", "2026-08-26", "2026-08-30", ""],
            ["break", "Winter break", "2026-12-19", "2027-01-04", ""],
            ["term", "Term 2", "2027-01-05", "2027-04-02", ""],
            ["exam-week", "IGCSE / AS / A2", "2027-04-05", "2027-04-30", "External exams"],
            ["term", "Term 3", "2027-04-25", "2027-06-25", ""],
        ],
        dropdowns={1: ["term","holiday","break","exam-week"]},
    )

    # ---- 6. LESSON PLANS (OPTIONAL) ----
    add_template(wb, "6. Lesson plans (OPTIONAL)",
        "OPTIONAL · One row per scheduled lesson with learning objectives. Optional for the initial demo; "
        "powers the classroom recap features in V2.",
        columns=[
            ("Section code", 14, "e.g. 9A, 10B"),
            ("Subject code", 14, "Match Manhaj subject codes (Ph, Bi, Ma, En, Ar, etc.)"),
            ("Date", 14, "YYYY-MM-DD"),
            ("Period #", 10, ""),
            ("Teacher name", 24, ""),
            ("Topic", 36, ""),
            ("Learning objective", 50, "What should the student be able to do by the end?"),
            ("Resources", 30, "Textbook chapter, worksheet, video link"),
        ],
        samples=[
            ["10A", "Ch", "2026-09-08", 3, "Sandra Swart", "Equilibrium · Le Chatelier", "Predict shift direction given pressure / temperature / concentration changes", "IGCSE Chem Ch. 22 + worksheet"],
            ["9B", "En", "2026-09-08", 1, "Marli Shaw", "Persuasive essay structure", "Identify and apply PEEL paragraph structure in a 200-word response", "Style guide handout"],
        ],
    )

    # ---- 7. ASSESSMENTS (OPTIONAL) ----
    add_template(wb, "7. Assessments (OPTIONAL)",
        "OPTIONAL · Past assessment scores. Drives the grade-trend, percentile, and rubric-anchor signals. "
        "If unavailable, demo uses Manhaj-generated illustrative dataset (clearly labelled).",
        columns=[
            ("Student ID", 16, ""),
            ("Section code", 14, ""),
            ("Subject code", 14, ""),
            ("Assessment label", 28, "e.g. 'Term 1 quiz 3', 'Mock paper 1'"),
            ("Type", 16, "quiz / test / exam / project / homework"),
            ("Date", 14, "YYYY-MM-DD"),
            ("Score", 10, "Numeric"),
            ("Max score", 10, "Numeric"),
            ("Weight", 10, "0.0–1.0 — weight in term grade"),
            ("Notes", 24, ""),
        ],
        samples=[
            ["ISO-2843", "10A", "Ch", "Term 1 quiz 1", "quiz", "2026-09-25", 18, 20, 0.1, ""],
            ["ISO-2843", "10A", "Ma", "Algebra mock", "test", "2026-10-12", 76, 100, 0.3, ""],
        ],
        dropdowns={5: ["quiz","test","exam","project","homework"]},
    )

    # ---- 8. CHART OF ACCOUNTS (OPTIONAL) ----
    add_template(wb, "8. Chart of accounts (OPT)",
        "OPTIONAL · Finance module only. Send during the Finance pilot conversation — not needed for the initial Admin + Classroom demo.",
        columns=[
            ("Account code", 14, "School's own GL code"),
            ("Account name", 36, ""),
            ("Category", 18, "revenue / payroll / operations / capex / other"),
            ("Parent account", 18, "Optional · for hierarchical CoA"),
            ("Active", 10, "Y / N"),
        ],
        samples=[
            ["4000", "Tuition revenue · KG-G5", "revenue", "4000-group", "Y"],
            ["6000", "Teaching staff payroll", "payroll", "6000-group", "Y"],
            ["7100", "Utilities · electricity", "operations", "7000-group", "Y"],
        ],
        dropdowns={3: ["revenue","payroll","operations","capex","other"], 5: ["Y","N"]},
    )

    # ---- 9. ALUMNI PLACEMENTS (OPTIONAL) ----
    add_template(wb, "9. Alumni placements (OPT)",
        "OPTIONAL but HIGH-VALUE · Last 5 years of where graduating students went. "
        "Unlocks the university-fit signal in monthly parent reports. Without this we can demo but cannot anchor to ISO's own history.",
        columns=[
            ("Graduation year", 14, ""),
            ("Student ID (or anonymous)", 22, "If sensitive, use 'ANON-{year}-{n}'"),
            ("University", 36, ""),
            ("Country", 14, ""),
            ("Programme", 28, "Degree + major"),
            ("IGCSE / AS / A2 grade summary", 30, "Optional · e.g. 'A*A*ABB at A2'"),
            ("Notes", 30, ""),
        ],
        samples=[
            [2024, "ANON-2024-01", "Sultan Qaboos University", "Oman", "BSc Medicine", "A*AAAB", "Pre-med pathway"],
            [2024, "ANON-2024-02", "American University of Beirut", "Lebanon", "BBA Finance", "AABB", ""],
            [2023, "ANON-2023-01", "University of Manchester", "UK", "BEng Chemical Engineering", "A*A*A", ""],
        ],
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
