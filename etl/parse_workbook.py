"""
ETL: parses the school's 24-25 provisional workbook into structured JSON for the demo.

NOTE ON SOURCE FILE LOCATION (post 2026-05-25 storage policy):
  The xlsx referenced below is NOT committed to git. See docs/data_storage_policy.md.
  - For local development: place the file at ../data/source/24-25 provisional_18-8-2024.xlsx
    (the data/source/ folder is gitignored).
  - At Tier 1 (Supabase live): this script becomes the *local-dev fallback*. The production
    path downloads the file from Supabase Storage `raw-uploads/{school_id}/...` and writes
    directly into Postgres instead of JSON files.

NOTE ON HARD-CODED MAPPING:
  This script encodes ISO's specific 26-27A sheet layout (column groupings, subject-code
  normalisation, etc.). When we onboard school #2, this hard-coding gets refactored into a
  generic config-driven runner — see docs/universal_mapper_spec.md. The ISO logic in this
  file becomes the first row in `source_mapping_configs`.

Outputs (all in ../data/processed/, also gitignored from Tier 1 onward):
  teachers.json           — [{id, name, dept, primary_subject, cap, assigned}]
  subjects.json           — [{code, name_en, name_ar, department, is_ap, is_self_study}]
  sections.json           — [{code, grade_level, label, stream}]
  load_matrix.json        — [{teacher_id, section_code, subject_code, weekly_periods}]
  course_offerings.json   — from circulars (compulsory + elective bundles per grade)

Designed to be re-runnable.
"""
from __future__ import annotations
import json, re, hashlib
from pathlib import Path
from collections import defaultdict
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC  = ROOT / "data" / "source" / "24-25 provisional_18-8-2024.xlsx"
OUT  = ROOT / "data" / "processed"
OUT.mkdir(parents=True, exist_ok=True)

# Subject code → human label mapping. Codes come from "Faculty" sheet row 2.
# Names come from the course-selection circulars.
SUBJECT_CATALOG = {
    "A3":   {"name_en": "Arabic (Advanced)",        "name_ar": "اللغة العربية متقدم", "department": "Arabic"},
    "Ar":   {"name_en": "Arabic",                   "name_ar": "اللغة العربية",        "department": "Arabic"},
    "En":   {"name_en": "English",                  "name_ar": "اللغة الإنجليزية",     "department": "English"},
    "ER":   {"name_en": "English Reading",          "name_ar": "القراءة الإنجليزية",   "department": "English"},
    "ES":   {"name_en": "English Support",          "name_ar": "دعم اللغة الإنجليزية", "department": "English"},
    "F2":   {"name_en": "French (Stage 2)",         "name_ar": "اللغة الفرنسية ٢",     "department": "French"},
    "F3":   {"name_en": "French (Stage 3)",         "name_ar": "اللغة الفرنسية ٣",     "department": "French"},
    "SSE":  {"name_en": "Social Studies (English)", "name_ar": "الدراسات الاجتماعية (إنجليزي)", "department": "Social-English"},
    "Hi":   {"name_en": "History",                  "name_ar": "التاريخ",              "department": "Social-English"},
    "CV":   {"name_en": "Civics",                   "name_ar": "التربية الوطنية",      "department": "Social-English"},
    "dv":   {"name_en": "Environmental Management",  "name_ar": "الإدارة البيئية",      "department": "Social-English"},
    "Ec":   {"name_en": "Economics",                "name_ar": "الاقتصاد",             "department": "Social-English"},
    "BS":   {"name_en": "Business Studies",         "name_ar": "إدارة الأعمال",         "department": "Social-English"},
    "SSA":  {"name_en": "Arabic Social Studies",    "name_ar": "الدراسات الاجتماعية",   "department": "Social-Arabic"},
    "IS":   {"name_en": "Islamic Studies",          "name_ar": "التربية الإسلامية",     "department": "Social-Arabic"},
    "Sc":   {"name_en": "Science",                  "name_ar": "العلوم",                "department": "Science"},
    "Bi":   {"name_en": "Biology",                  "name_ar": "الأحياء",               "department": "Science"},
    "Bi AP":{"name_en": "Biology (Advanced)",       "name_ar": "الأحياء متقدم",         "department": "Science", "is_ap": True},
    "Bi SS":{"name_en": "Biology (Self-Study)",     "name_ar": "الأحياء دراسة ذاتية",   "department": "Science", "is_self_study": True},
    "Ch":   {"name_en": "Chemistry",                "name_ar": "الكيمياء",              "department": "Science"},
    "Ch AP":{"name_en": "Chemistry (Advanced)",     "name_ar": "الكيمياء متقدم",        "department": "Science", "is_ap": True},
    "Ph":   {"name_en": "Physics",                  "name_ar": "الفيزياء",              "department": "Science"},
    "Ph AP":{"name_en": "Physics (Advanced)",       "name_ar": "الفيزياء متقدم",        "department": "Science", "is_ap": True},
    "IT":   {"name_en": "ICT",                      "name_ar": "تكنولوجيا المعلومات",   "department": "Science"},
    "Ma":   {"name_en": "Mathematics",              "name_ar": "الرياضيات",             "department": "Math"},
    "Ma AP":{"name_en": "Mathematics (Advanced)",   "name_ar": "الرياضيات متقدم",       "department": "Math", "is_ap": True},
    "MS":   {"name_en": "Math Support",             "name_ar": "دعم الرياضيات",         "department": "Math"},
    "Mu":   {"name_en": "Music",                    "name_ar": "الموسيقى",              "department": "Recreational"},
    "PE":   {"name_en": "Physical Education",       "name_ar": "التربية البدنية",       "department": "Recreational"},
    "rt":   {"name_en": "Art",                      "name_ar": "الفن",                  "department": "Recreational"},
    "Ex":   {"name_en": "Examinations",             "name_ar": "الامتحانات",            "department": "Assessment"},
    "lb":   {"name_en": "Library / Lab",            "name_ar": "المكتبة / المختبر",     "department": "Assessment"},
}

# Map the short codes seen in the 26-27A sub-columns (En/Ma/Mu/rt/etc.) to the Faculty codes.
# The sheet is inconsistent about case + 2- vs 3-letter abbreviations across years; we normalise.
SHORT_TO_CODE = {
    # Canonical (Faculty sheet) codes
    "En": "En", "Ma": "Ma", "Mu": "Mu", "rt": "rt", "PE": "PE",
    "Ar": "Ar", "A3": "A3", "IS": "IS", "CV": "CV",
    "Sc": "Sc", "Bi": "Bi", "Ch": "Ch", "Ph": "Ph", "IT": "IT",
    "Hi": "Hi", "Ec": "Ec", "BS": "BS", "SSE": "SSE", "SSA": "SSA",
    "F2": "F2", "F3": "F3", "MS": "MS", "ES": "ES", "ER": "ER",
    "Ex": "Ex", "lb": "lb", "dv": "dv",
    # Uppercase / extended variants observed in 26-27A
    "MA": "Ma", "M1": "Ma",
    "EN": "En",
    "AR": "Ar", "AR3": "A3", "A2": "A3",
    "MU": "Mu", "RT": "rt", "AD": "rt",                   # AD = "Art & Design"
    "SC": "Sc", "BIO": "Bi", "BSS": "Bi SS",
    "CH": "Ch", "PH": "Ph", "PHY": "Ph",
    "IC": "IT",
    "HI": "Hi", "EC": "Ec", "ECO": "Ec", "EM": "dv",
    "ASS": "SSA", "S2": "SSE",
    "FR3": "F3",
    "EX": "Ex", "LB": "lb",
    "GP": "Ex", "WB": "lb",                                # GP = group/exam prep; WB = wellbeing/library — best-guess
    "ML": "Ar", "R9": "Ar",                                # ML = mother-language; R9 = recitation grade-9
}

def parse_section_code(code: str) -> dict:
    """KG1A → {grade='KG1', label='A', stream='regular'}
       11 AS → {grade='11',  label='AS', stream='AS'}
       1-2 AL → {grade='1-2', label='AL', stream='AL_combined'}"""
    code = code.strip()
    if code.startswith("KG"):
        return {"grade_level": code[:3], "label": code[3:], "stream": "regular"}
    m = re.match(r"^([0-9]+(?:-[0-9]+)?)\s*(.*)$", code)
    if not m:
        return {"grade_level": code, "label": "", "stream": "regular"}
    grade, suffix = m.group(1), m.group(2).strip()
    stream = "regular"
    if suffix == "AS":   stream = "AS"
    elif suffix == "A2": stream = "A2"
    elif suffix == "AL": stream = "AL_combined"
    elif "-" in grade and suffix == "AL": stream = "AL_combined"
    return {"grade_level": grade, "label": suffix, "stream": stream}


def parse_26_27a(wb):
    """Walks the load-matrix sheet and emits teacher + load rows.

    Layout discovered: row 1 has section codes at sparse columns. Each section
    occupies a contiguous column block until the next section header.
    Each teacher occupies a multi-row block; the row containing the integer
    in col 0 ("S#") is the load row. The row containing English/Math/etc.
    short codes (1-3 chars, no digits) is the subject-header row for that
    teacher's section block.
    """
    ws = wb["26-27A"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    width = len(hdr)

    # 1. Section column ranges
    section_starts = [(i, str(hdr[i]).strip()) for i in range(5, width) if hdr[i] is not None and str(hdr[i]).strip()]
    sections = []
    for k, (col, code) in enumerate(section_starts):
        end = section_starts[k+1][0] if k+1 < len(section_starts) else width
        sections.append({"code": code, "col_start": col, "col_end": end, **parse_section_code(code)})

    # 2. Teacher rows. A teacher row has integer in col 0 and a string in col 1.
    teachers = []
    teacher_row_idx = []
    for i, r in enumerate(rows):
        if r[0] is None or r[1] is None: continue
        try:
            sn = int(r[0])
        except (ValueError, TypeError):
            continue
        name = str(r[1]).strip()
        if not name: continue
        cap = int(r[2]) if r[2] is not None else 30
        primary_subject = str(r[3]).strip() if r[3] else ""
        assigned = int(r[4]) if isinstance(r[4], (int, float)) else 0
        tid = f"T{sn:03d}"
        teachers.append({
            "id": tid, "serial": sn, "name": name, "cap": cap,
            "primary_subject": primary_subject, "assigned_in_sheet": assigned,
        })
        teacher_row_idx.append((tid, i))

    # 3. Load matrix. For each teacher row, scan its section column blocks.
    #    The subject sub-header sits in a row close-by (search ±3 rows for short alpha codes).
    load = []
    for ti, (tid, trow) in enumerate(teacher_row_idx):
        # Determine block end (start of next teacher row, or sheet end)
        next_trow = teacher_row_idx[ti+1][1] if ti+1 < len(teacher_row_idx) else len(rows)
        # Within this block, find the subject-header row per section
        for sec in sections:
            # Find header row: scan trow..next_trow for short alpha values in this section's columns
            header_row = None
            for rr in range(trow, min(next_trow, trow+4)):
                cells = [rows[rr][c] for c in range(sec["col_start"], sec["col_end"])]
                alpha_like = sum(1 for c in cells if isinstance(c, str) and 1 <= len(c.strip()) <= 4 and not c.strip().isdigit())
                if alpha_like >= 1:
                    header_row = rr
                    break
            # Load values are on the teacher's own row at the section columns
            for c in range(sec["col_start"], sec["col_end"]):
                val = rows[trow][c]
                if val is None: continue
                if not isinstance(val, (int, float)) or val <= 0: continue
                # Subject code from header_row at the same column
                short = None
                if header_row is not None:
                    h = rows[header_row][c]
                    if h is not None and isinstance(h, str):
                        short = h.strip()
                code = SHORT_TO_CODE.get(short, short) if short else None
                if not code:
                    code = "?"
                load.append({
                    "teacher_id": tid,
                    "section_code": sec["code"],
                    "subject_code": code,
                    "subject_short_raw": short,
                    "weekly_periods": int(val),
                    "source_cell": f"26-27A!R{trow+1}C{c+1}",
                })

    return teachers, sections, load


def parse_faculty(wb):
    """Cleaner cross-tab. Used to enrich teacher subject coverage."""
    ws = wb["Faculty"]
    rows = list(ws.iter_rows(values_only=True))
    # Row 2 has subject codes; row 1 has department groups
    subj_codes = []
    for i, v in enumerate(rows[1]):
        if v is not None and str(v).strip():
            subj_codes.append((i, str(v).strip()))

    crosstab = []
    for r in rows[2:]:
        if r[0] is None: continue
        name = str(r[0]).strip()
        if not name: continue
        for col, code in subj_codes:
            v = r[col] if col < len(r) else None
            if isinstance(v, (int, float)) and v > 0:
                crosstab.append({"teacher_name": name, "subject_code": code, "weekly_periods": int(v)})
    return crosstab


def course_offerings():
    """Hand-coded from the four circulars (.doc files). One source of truth for
    G9–G12 compulsory + elective bundles. KG-G8 not specified in circulars; we
    infer from the load-matrix sub-columns at demo time."""
    return {
        "9": {
            "stage": "IGCSE",
            "pick_count": 5,
            "compulsory": ["Ma", "En", "Ar", "SSA", "IS"],
            "language_alt": {"Ar": ["F2", "F3"]},
            "bundles": [
                {"label": "Science 1", "options": ["Ph", "BS"]},
                {"label": "Science 2", "options": ["Ch", "dv"]},
                {"label": "Science 3", "options": ["Bi", "IT"]},
                {"label": "Humanities/Arts", "options": ["Ec", "Hi", "rt"]},
                {"label": "Activity", "options": ["PE", "rt"]},
            ],
        },
        "10": {
            "stage": "IGCSE",
            "pick_count": 5,
            "compulsory": ["Ma", "En", "Ar", "SSA", "IS"],
            "language_alt": {"Ar": ["F2", "F3"]},
            "bundles": [
                {"label": "Science 1", "options": ["Ph", "Bi"]},
                {"label": "Science 2", "options": ["Ch", "dv"]},
                {"label": "Applied/Arts", "options": ["IT", "BS", "rt"]},
                {"label": "Humanities", "options": ["Hi", "Bi SS", "Ec"]},
                {"label": "Activity", "options": ["PE", "rt"]},
            ],
        },
        "11": {
            "stage": "AS",
            "pick_count": 3,
            "compulsory": ["Ma", "En", "Ar", "IS", "CV"],
            "language_alt": {},
            "bundles": [
                {"label": "Science/Business 1", "options": ["Ph", "BS"]},
                {"label": "Science/Business 2", "options": ["Ch", "Ec"]},
                {"label": "Bio/IT", "options": ["Bi", "IT"]},
                {"label": "Activity", "options": ["PE", "rt"]},
            ],
        },
        "12": {
            "stage": "A2",
            "pick_count": 3,
            "compulsory": ["Ma", "En", "CV", "IS"],
            "language_alt": {"Ar": ["F2", "F3"]},
            "bundles": [
                {"label": "Science 1", "options": ["Ph", "Bi"]},
                {"label": "Science/Business", "options": ["Ch", "Ec"]},
                {"label": "Applied", "options": ["BS", "Bi SS", "IT"]},
                {"label": "Activity", "options": ["rt", "PE"]},
            ],
        },
    }


def file_sha256(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""): h.update(chunk)
    return h.hexdigest()


def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    teachers, sections, load = parse_26_27a(wb)
    faculty_crosstab = parse_faculty(wb)

    # Compute assigned per teacher from the load rows (cross-check vs sheet's T#P)
    computed = defaultdict(int)
    for r in load:
        computed[r["teacher_id"]] += r["weekly_periods"]
    for t in teachers:
        t["assigned_computed"] = computed.get(t["id"], 0)
        t["slack"] = t["cap"] - t["assigned_computed"]

    # Subjects list
    subjects_seen = sorted(set(r["subject_code"] for r in load) | set(r["subject_code"] for r in faculty_crosstab))
    subjects = []
    for code in subjects_seen:
        if code == "?": continue
        info = SUBJECT_CATALOG.get(code, {"name_en": code, "name_ar": "", "department": "Unknown"})
        subjects.append({
            "code": code, "name_en": info["name_en"], "name_ar": info.get("name_ar", ""),
            "department": info["department"], "is_ap": info.get("is_ap", False),
            "is_self_study": info.get("is_self_study", False),
        })

    manifest = {
        "source_file": SRC.name,
        "source_sha256": file_sha256(SRC),
        "school": {"name": "International School of Oman", "city": "Muscat", "country": "OM"},
        "academic_year": "2026-2027",
        "counts": {
            "teachers": len(teachers),
            "sections": len(sections),
            "subjects": len(subjects),
            "load_rows": len(load),
            "faculty_rows": len(faculty_crosstab),
        }
    }

    (OUT / "manifest.json").write_text(json.dumps(manifest, indent=2, ensure_ascii=False))
    (OUT / "teachers.json").write_text(json.dumps(teachers, indent=2, ensure_ascii=False))
    (OUT / "sections.json").write_text(json.dumps(sections, indent=2, ensure_ascii=False))
    (OUT / "subjects.json").write_text(json.dumps(subjects, indent=2, ensure_ascii=False))
    (OUT / "load_matrix.json").write_text(json.dumps(load, indent=2, ensure_ascii=False))
    (OUT / "faculty_crosstab.json").write_text(json.dumps(faculty_crosstab, indent=2, ensure_ascii=False))
    (OUT / "course_offerings.json").write_text(json.dumps(course_offerings(), indent=2, ensure_ascii=False))

    print(json.dumps(manifest, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
